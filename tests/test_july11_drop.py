"""July 11 drop is frozen as A/B input, not as judged evidence."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DROP = ROOT / "data/july11_gallery_4136050"
PROVENANCE = json.loads((DROP / "provenance.json").read_text())
MANIFEST = json.loads((DROP / "manifest.json").read_text())
SHEET = DROP / "BlueToad_2026-07-11_BidSheet.xlsx"


def test_ad_hoc_bidsheet_matches_pinned_sha256():
    expected = PROVENANCE["ad_hoc_bidsheet"]
    digest = hashlib.sha256(SHEET.read_bytes()).hexdigest()
    assert SHEET.stat().st_size == expected["byte_size"]
    assert digest == expected["sha256"]


def test_ad_hoc_bidsheet_has_the_desktop_tabs():
    wb = load_workbook(SHEET, read_only=True, data_only=True)
    assert wb.sheetnames == PROVENANCE["ad_hoc_bidsheet"]["sheets"]
    auction = {row[0]: row[1] for row in wb["Auction Info"].iter_rows(values_only=True) if row[0]}
    assert auction["Date"].startswith("Saturday July 11, 2026")
    assert "4136050" in str(auction["Listing"])
    training = wb["All Lots (Training)"]
    assert training.max_row == 453  # header + 452 photos
    bids = wb["Bid Sheet"]
    items = sum(
        1
        for row in bids.iter_rows(min_row=2, values_only=True)
        if row[0] in {"A", "B", "C"}
    )
    assert items == 88


def test_manifest_is_the_452_photo_drop():
    assert MANIFEST["listing_id"] == "4136050"
    assert MANIFEST["total_photos"] == 452
    assert MANIFEST["captioned_photos"] == 324
    assert len(MANIFEST["photos"]) == 452
    sequences = [p["sequence"] for p in MANIFEST["photos"]]
    assert sequences == sorted(sequences)
    assert len(set(p["photo_id"] for p in MANIFEST["photos"])) == 452


def test_cached_images_hash_to_the_manifest_when_present():
    photos = MANIFEST["photos"]
    present = [
        p for p in photos
        if (ROOT / p["local_path"]).is_file() and (ROOT / p["local_path"]).stat().st_size > 0
    ]
    if not present:
        pytest.skip("july11 image cache is not on this clone")
    assert len(present) == 452
    bad = []
    for photo in present:
        data = (ROOT / photo["local_path"]).read_bytes()
        digest = hashlib.sha256(data).hexdigest()
        if photo.get("download_status") != "usable" or photo.get("sha256") != digest:
            bad.append(photo["filename"])
    assert bad == []


def test_drop_readme_does_not_treat_the_scorecard_as_the_ab():
    text = (DROP / "README.md").read_text()
    assert "not submission evidence" in text.lower()
    assert "Side A" in text
    assert "run_july11_benchmark.py" in text
    assert PROVENANCE["submission_evidence"] is False
