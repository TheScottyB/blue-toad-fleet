#!/usr/bin/env python3
"""
scripts/run_vertex_pipeline.py — End-to-End Live Vertex AI Sourcing Pipeline.

Executes the two-stage model pipeline:
1. Photo Intake & evidence-gated grouping
2. Stage 1 Triage (gemini-3.5-flash-lite / structured filtering)
3. Stage 2 Appraisal (gemini-3.6-flash live on Vertex AI for candidates)
4. Question Queue resolution with memory rules (build_queue, StandingRule)
5. Pure Bid Math allocation under the configured cap, fee, and bid increments
6. Compiles final absentee bid email and Excel sheet.
"""

import json
import os
import shutil
import sys
import argparse
from dataclasses import asdict, dataclass, field, replace
from pathlib import Path
from types import MappingProxyType
from typing import Callable, Mapping

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.intake.embed import load_reshoot_edges, sha256_file
from src.intake.manifest import parse_drop, TriagedPhoto
from src.intake.spatial import (
    SpatiallyTaggedPhoto, Zone, SurfaceSignature, apply_trajectory,
    load_observations,
)
from src.intake.puzzle import as_lot_groups, puzzle_loop, walk_proposal_edges
from src.appraisal import (
    Appraisal, Confidence as AppConfidence, Question, QuestionKind,
    StandingRule, build_queue
)
from src.appraiser import AppraisalEngine
from src.appraiser.containers import append_visible_contents, price_container_evidence
from src.appraiser.grounded_batch import (
    GroundedPricingPipeline, grounded_reference_comps, grounded_status_reason,
)
from src.appraiser.pricing import MIN_CALLS
from src.appraiser.routing import (
    APPRAISAL_MODEL, TRIAGE_MODEL, estimate_cost_usd,
    rate_snapshot_usd_per_million,
)
from src.evidence.telemetry import UsageTelemetry
from src.evidence import load_absorption_evidence
from src.assemble.email import compile_absentee_email
from src.bidmath import (
    Lot, CompEstimate, Confidence as BidConfidence, Priority, Decision,
    price_lot, allocate, summarize, ABSENTEE_FEE, snap_to_increment,
    mechanic_from_ruling, BidMechanic, units_committed, opening_bid, is_choice_lot,
    elect,
)

# Reference valuation comps for approved candidate categories (matching shop pricing bands)
REFERENCE_COMPS = {
    "BT-001": {"low": 250.0, "high": 320.0, "sources": 4, "conf": BidConfidence.HIGH, "cat": "vintage cards", "desc": "Vintage Topps Baseball Cards (1959-69 Golden Era Stars)"},
    "BT-041": {"low": 100.0, "high": 130.0, "sources": 3, "conf": BidConfidence.HIGH, "cat": "phonograph / records", "desc": "Edison rolls (11-12 canisters + bare roll)"},
    "BT-002": {"low": 65.0, "high": 75.0, "sources": 3, "conf": BidConfidence.HIGH, "cat": "jewelry", "desc": "Estate Costume Jewelry (Tray 12/14/16: 50-70 pcs)"},
    "BT-087": {"low": 65.0, "high": 75.0, "sources": 3, "conf": BidConfidence.HIGH, "cat": "jewelry", "desc": "costume jewelry (Tray Lot 2)"},
    "BT-181": {"low": 65.0, "high": 75.0, "sources": 3, "conf": BidConfidence.HIGH, "cat": "jewelry", "desc": "estate costume jewelry (Tray Lot 3)"},
    "BT-050": {"low": 65.0, "high": 75.0, "sources": 3, "conf": BidConfidence.HIGH, "cat": "vintage toys", "desc": "Lionel building set"},
    "BT-021": {"low": 52.0, "high": 62.0, "sources": 2, "conf": BidConfidence.HIGH, "cat": "vintage electronics", "desc": "princess phone"},
    "BT-048": {"low": 52.0, "high": 62.0, "sources": 2, "conf": BidConfidence.HIGH, "cat": "vintage smalls", "desc": "ET nightlight"},
    "BT-235": {"low": 38.0, "high": 48.0, "sources": 2, "conf": BidConfidence.HIGH, "cat": "advertising / bottles", "desc": "Century Progress bottle"},
    "BT-016": {"low": 38.0, "high": 48.0, "sources": 2, "conf": BidConfidence.HIGH, "cat": "vintage cards", "desc": "trading cards"},
    "BT-030": {"low": 38.0, "high": 48.0, "sources": 2, "conf": BidConfidence.HIGH, "cat": "vintage cards", "desc": "non-sport trading cards"},
    "BT-066": {"low": 26.0, "high": 32.0, "sources": 2, "conf": BidConfidence.HIGH, "cat": "vintage toys", "desc": "hand held video games (5 Radica/LCD units)"},
}

# Decisions the owner made in the 2026-08-20 collaborative chat, kept next to the
# comps they apply to. These are answers, not appraisals.
#
# The record is partial and says so. He remembers trimming the card lots to the
# best one and capping the jewellery at $25; the rest of that conversation is not
# recoverable. Anything not recorded here follows the appraiser, which is the
# honest default — a decision nobody can produce is not a decision.
#
# The appraiser scores a lot on how well it fits the eight categories it was told
# the shop buys. It scored the bulk costume jewelry and the junk-wax card box at
# 0.20 and asked whether the shop wanted them at all — a fair question, and one
# the owner answered: he takes them for the storefront. `fit` carries his answer,
# `why` carries the reason the fit score could not see.
#
# Recorded here rather than written over the appraisal, so the console can show
# both what the model concluded and what the owner decided.
OPERATOR_APPROVED = {
    # fit: what the owner decided about category fit. None means he declined the
    #      lot outright, and the sheet must not bid it.
    # cap: a max bid he set. In absentee proxy bidding a max is a ceiling you
    #      only reach if someone pushes you there, so it is what he is willing
    #      to go to, not what he expects to pay. A condition penalty must not
    #      walk it down — bidding under a defensive cap loses the lot to the
    #      next bidder, which is the outcome the cap existed to prevent.
    "BT-001": {"fit": 0.90, "cap": 100.00,
               "why": "collab: top of the three card lots, $100 cap agreed"},
    "BT-016": {"fit": None,
               "why": "collab: 'give me the top 1 of the three card lots' — BT-001 took it"},
    "BT-030": {"fit": None,
               "why": "collab: 'give me the top 1 of the three card lots' — BT-001 took it"},

    "BT-002": {"fit": 0.90, "cap": 25.00,
               "why": "collab: buys bulk estate costume jewelry, max bid $25",
               # The auctioneer's own ruling on the lot_grouping question the
               # appraiser raised: "Is the auction bid for a single tray (12, 14
               # or 16) or for all trays shown together?" Bill Theesfield by
               # email 2026-08-21: "Yes, that is a x3 bid." Recorded as words so
               # the sheet derives the money from the ruling rather than someone
               # retyping $75 into an email under the cutoff.
               "ruling": "take all three trays at x3"},
    # REVISED DOWN 2026-08-21. The absentee sheet that actually went to Blue
    # Toad reads "START $5.00   MAX $15.00   (revised down from $25.00)". This
    # table still said $25.00, and apply_operator_cap sets max_bid to the cap
    # unconditionally by design — it is not a ceiling that clamps down — so the
    # bid went out $10 ABOVE what the operator authorised, on his own cash.
    # Substituting 15.00 reproduces the sent sheet's footer exactly:
    # $275.00 committed / $316.25 all-in.
    "BT-087": {"fit": 0.90, "cap": 15.00,
               "why": "collab: bulk estate costume jewelry; revised down to $15 "
                      "on the sent 2026-08-21 sheet"},
    # DUPLICATE of BT-002, confirmed three independent ways on 2026-08-21:
    #  - visual: seq 181 is a close-up of trays 12 and 14 from seq 2. The gold
    #    flat-link mesh necklace, the cream oval-bead strand, the coin-charm
    #    bracelet with the starfish, the blue lapis-glass pendant and the green
    #    enamel Christmas tree brooch are the same objects on the same velvet
    #    tray on the same concrete floor; tray 14's bead necklaces sit at 181's
    #    right margin.
    #  - gemini-embedding-2 ranks seq 2 as seq 181's #1 neighbour, cos 0.906,
    #    where dHash ranks it #94 — adjacency-based grouping cannot see it
    #    because the auctioneer returned to the table 179 frames later.
    #  - the revised absentee sheet that actually went to Blue Toad drops it.
    # BT-002 already commits all three trays at x3, so bidding this too buys the
    # same trays twice. Declined rather than deleted, so the reason survives.
    "BT-181": {"fit": None,
               "why": "duplicate of BT-002 (close-up of trays 12/14) — "
                      "BT-002 takes all three at x3"},

    "BT-021": {"fit": 0.90, "why": "collab: vintage telephones sell in store"},
    "BT-041": {"fit": 0.90, "why": "collab: Edison cylinders are an alpha pick this cycle"},
    "BT-048": {"fit": 0.90, "why": "collab: licensed 80s character pieces move fast"},
    "BT-050": {"fit": 0.90, "why": "collab: Lionel set approved"},
    "BT-066": {"fit": 0.90, "why": "collab: handheld electronic games approved"},
    "BT-235": {"fit": 0.90, "why": "collab: Century of Progress bottle approved"},
}


@dataclass(frozen=True)
class PipelineConfig:
    cycle_id: str
    listing_id: str
    data_dir: Path
    output_dir: Path | None
    budget_cap: float
    auto_send_threshold: float
    auction_title: str
    auction_date: str
    auction_timezone: str
    auction_deadline: str
    venue: str
    email_to: str = "info@bluetoadauctions.com"
    force_live_vertex: bool = False
    reference_comps: Mapping = field(default_factory=lambda: MappingProxyType({}))
    operator_approved: Mapping = field(default_factory=lambda: MappingProxyType({}))
    standing_rules: tuple[StandingRule, ...] = ()
    cycle_questions: tuple[Question, ...] = ()
    enable_grounded_pricing: bool = False
    pricing_min_fit: float = 0.70
    pricing_workers: int = 6
    pricing_engine_factory: Callable | None = None

    def __post_init__(self) -> None:
        object.__setattr__(self, "data_dir", Path(self.data_dir))
        if self.output_dir is not None:
            object.__setattr__(self, "output_dir", Path(self.output_dir))
        if not all(str(value or "").strip() for value in (
            self.cycle_id, self.listing_id, self.auction_title, self.auction_date,
            self.auction_timezone, self.auction_deadline, self.venue, self.email_to,
        )):
            raise ValueError("pipeline configuration is missing required metadata")
        if self.budget_cap <= 0 or not 0 <= self.auto_send_threshold <= self.budget_cap:
            raise ValueError("pipeline budget or auto-send threshold is invalid")
        if self.pricing_workers < 1:
            raise ValueError("pricing_workers must be positive")


@dataclass(frozen=True)
class PipelineResult:
    photos: tuple[dict, ...]
    lot_groups: tuple
    lots: tuple[Lot, ...]
    decisions: tuple[Decision, ...]
    summary: object
    queue: object
    captions: Mapping[str, str]
    pipeline_state_path: Path
    bid_sheet_path: Path
    email_path: Path


def execute_pipeline(config: PipelineConfig) -> PipelineResult:
    """Canonical typed boundary; ``run_pipeline`` remains a compatibility shim."""
    return run_pipeline(
        cycle_id=config.cycle_id,
        listing_id=config.listing_id,
        data_dir=str(config.data_dir),
        output_dir=str(config.output_dir) if config.output_dir is not None else None,
        budget_cap=config.budget_cap,
        auto_send_threshold=config.auto_send_threshold,
        force_live_vertex=config.force_live_vertex,
        auction_title=config.auction_title,
        auction_date=config.auction_date,
        auction_timezone=config.auction_timezone,
        auction_deadline=config.auction_deadline,
        venue=config.venue,
        email_to=config.email_to,
        reference_comps=dict(config.reference_comps),
        operator_approved=dict(config.operator_approved),
        standing_rules=config.standing_rules,
        cycle_questions=config.cycle_questions,
        enable_grounded_pricing=config.enable_grounded_pricing,
        pricing_min_fit=config.pricing_min_fit,
        pricing_workers=config.pricing_workers,
        pricing_engine_factory=config.pricing_engine_factory,
    )

def trusted_lot_flags(verdict, caption: str, previous_captioned: bool,
                      index: int) -> tuple[bool, bool]:
    """
    (is_lot, same_lot_as_previous) — how far Stage 1 is believed about boundaries.

    Triage is a cheap model looking at one photograph. It is good at "is this
    worth a slow look" and unreliable about where one lot ends and the next
    begins. On the first live corpus run it marked BT-181 — captioned "estate
    costume jewelry" by the house — as another angle of the necklaces two photos
    earlier. It was merged away, stopped being a lot, and a bid the owner had
    capped at $25 left the sheet without anything erroring.

    So a caption the auctioneer wrote outranks a merge the model guessed. This
    gallery publishes no lot numbers, which makes the caption the only boundary
    signal there is. Overriding a merge is not overriding everything: a triage
    verdict that says these are separate is left alone.

    With no verdict at all, the caption heuristic stands — an uncaptioned photo
    following a captioned one is another angle of it.
    """
    if verdict is None:
        is_extra_angle = (not caption.strip()) and index > 0 and previous_captioned
        return (not is_extra_angle), is_extra_angle

    is_lot = bool(verdict.get("is_lot", True))
    same = bool(verdict.get("same_lot_as_previous", False))
    if same and caption.strip():
        same = False
    return is_lot, same


def select_appraisal_candidates(
    triage_results: list[dict],
    photos: list[dict],
    always_include: set[str],
    category_hints: dict | None = None,
) -> list[dict]:
    """
    Which lots earn a slow, expensive look.

    Every photo is a candidate. Triage may still emit worth_appraising; it does
    not decide coverage. always_include remains accepted but is no longer a gate.
    """
    verdicts = {t.get("photo_id"): t for t in triage_results}
    category_hints = category_hints or {}
    out = []
    for p in sorted(photos, key=lambda x: x.get("sequence", 0)):
        lot_id = f"BT-{p.get('sequence', 0):03d}"
        verdict = verdicts.get(p.get("photo_id"))
        summary = (verdict or {}).get("summary", "")
        out.append({
            "lot_id": lot_id,
            "caption": p.get("caption") or summary or "(uncaptioned)",
            "category_hint": (category_hints.get(lot_id, {}).get("cat")
                              or (verdict or {}).get("category")),
            "local_path": p.get("local_path"),
        })
    return out


def select_decomposition_candidates(
    triage_results: list[dict],
    photos: list[dict],
    appraised_lot_ids: set[str] | None = None,
) -> list[dict]:
    """Select bounded collection lots for the spatial isolation pass.

    This consumes an explicit structured triage signal rather than guessing
    from plural words in captions.  A Spatial Room Graph may attach a trusted
    boundary to a manifest photo; otherwise the decomposer locates it before
    cropping.  Only lots continuing to appraisal need the extra pass.
    """
    verdicts = {t.get("photo_id"): t for t in triage_results}
    out = []
    for p in sorted(photos, key=lambda x: x.get("sequence", 0)):
        verdict = verdicts.get(p.get("photo_id")) or {}
        if not verdict.get("needs_decomposition"):
            continue
        lot_id = f"BT-{p.get('sequence', 0):03d}"
        if appraised_lot_ids is not None and lot_id not in appraised_lot_ids:
            continue
        candidate = {
            "lot_id": lot_id,
            "caption": p.get("caption") or verdict.get("summary") or "(uncaptioned)",
            "local_path": p.get("local_path"),
        }
        boundary = p.get("spatial_boundary") or p.get("container_boundary")
        if boundary is not None:
            candidate["spatial_boundary"] = boundary
        if p.get("spatial_context"):
            candidate["spatial_context"] = p["spatial_context"]
        if p.get("container_type"):
            candidate["container_type"] = p["container_type"]
        out.append(candidate)
    return out


def operator_lot_inputs(
    lot_id: str, raw_appraisal: dict, operator_approved: dict | None = None,
) -> tuple[float, float]:
    """
    (fit_score, condition_penalty) for a candidate lot.

    The owner's decision carries the fit — he knows things a category-fit score
    cannot, and a lot he declined comes back at 0.0 so the allocator skips it.
    The appraiser's condition reading is carried through untouched: nothing in
    the collaborative review was about visible damage.

    One function because the console and the pipeline used to build these
    separately, and a decision applied in one was invisible in the other.
    """
    decisions = OPERATOR_APPROVED if operator_approved is None else operator_approved
    decision = decisions.get(lot_id, {})
    fit = decision.get("fit", float(raw_appraisal.get("fit_score", 0.5)))
    penalty = float(raw_appraisal.get("condition_penalty", 0.0))
    return (0.0 if fit is None else float(fit)), penalty


def comp_from_reference(
    comp_info: Mapping,
    raw_appraisal: Mapping,
) -> tuple[CompEstimate, dict, str | None]:
    """Resolve a whole-lot comp or an explicit alpha-plus-bulk container record.

    Container component evidence is optional but typed: callers must supply a
    cited ``bulk_floor`` and may supply a cited ``alpha_comp``. The model's
    decomposition alone can never promote the alpha into the money path.
    """
    record = dict(comp_info)
    container = record.get("container_evidence")
    decomposition = raw_appraisal.get("container_decomposition")
    if isinstance(container, Mapping) and isinstance(decomposition, Mapping):
        resolved = price_container_evidence(
            decomposition,
            alpha_comp=container.get("alpha_comp"),
            bulk_floor=container.get("bulk_floor") or {},
        )
        record.update({
            "low": resolved.low,
            "high": resolved.high,
            "sources": resolved.source_count,
            "citations": list(resolved.citations),
            "provenance": "container_alpha_plus_bulk" if resolved.alpha_confirmed
                          else "container_bulk_floor",
        })
        confidence = (BidConfidence.HIGH if resolved.source_count >= 2
                      else BidConfidence.MEDIUM)
        return CompEstimate(
            low=resolved.low,
            high=resolved.high,
            source_count=resolved.source_count,
            confidence=confidence,
        ), record, resolved.upside_note
    return CompEstimate(
        low=record["low"],
        high=record["high"],
        source_count=record["sources"],
        confidence=record["conf"],
    ), record, None


def apply_operator_fit(lot):
    """Owner fit is keyed by the surviving lot_id, after reshoot merge.

    Stamping fit=None onto a member photo before merge lets a high-confidence
    close-up (BT-181) zero the surviving lot (BT-002). Apply after union.
    """
    lot_id = lot.lot_id.removeprefix("seq:")
    if lot_id != lot.lot_id:
        lot = replace(lot, lot_id=lot_id)
    fit, _ = operator_lot_inputs(lot_id, {
        "fit_score": lot.fit_score,
        "condition_penalty": lot.condition_penalty,
    })
    if fit == lot.fit_score:
        return lot
    return replace(lot, fit_score=fit)


def apply_operator_cap(decision, operator_approved: dict | None = None):
    """
    Use the max bid the owner set, where he set one.

    His number stands in both directions. "$100 defensive cap" on the Topps run
    means go to $100 rather than lose it; "move max bid down to 25" on the
    jewellery means 25. Treating either as a ceiling over a computed figure
    quietly bids less than he authorised, and in a proxy auction bidding less
    than your ceiling is how you lose the lot for five dollars.
    """
    decisions = OPERATOR_APPROVED if operator_approved is None else operator_approved
    cap = decisions.get(decision.lot_id, {}).get("cap")
    if cap is None or decision.max_bid is None or decision.max_bid == cap:
        return decision
    return replace(decision, max_bid=cap,
                   all_in=round(cap * (1.0 + ABSENTEE_FEE), 2))


DEFAULT_STANDING_RULES = [
    StandingRule(
        kind=QuestionKind.APPETITE,
        category="dinnerware / pottery",
        answer="SKIP — store has zero room for dishes or box sets",
        learned_cycle="2026-08-22",
    ),
    StandingRule(
        kind=QuestionKind.POLICY,
        category="sports memorabilia",
        answer="SKIP raw uncertified autographs — store currently overstocked",
        learned_cycle="2026-08-22",
    ),
    StandingRule(
        kind=QuestionKind.APPETITE,
        category="vintage tools",
        answer="SKIP — store backlog of unlisted tools",
        learned_cycle="2026-08-22",
    ),
    StandingRule(
        kind=QuestionKind.APPETITE,
        category="jewelry",
        answer="BUY — bulk estate costume jewelry moves in the storefront",
        learned_cycle="2026-08-20",
    ),
    StandingRule(
        kind=QuestionKind.APPETITE,
        category="vintage cards",
        answer="BUY — including junk-wax bulk boxes",
        learned_cycle="2026-08-20",
    ),
]


AUG22_DOMAIN_QUESTIONS = (
    Question(
        kind=QuestionKind.POLICY,
        category="sports memorabilia",
        prompt=("Uncertified Autographs (Jordan Hat BT-006, DiMaggio Hat BT-010): "
                "No visible PSA/JSA certificate in photos. Bid speculative raw "
                "floor or skip?"),
        lot_ids=("BT-006", "BT-010"),
        value_at_stake=800.0,
        confidence_gap=0.5,
    ),
    Question(
        kind=QuestionKind.APPETITE,
        category="dinnerware / pottery",
        prompt=("Under-Table Box Runs (Poppy Trail BT-073): Blue Toad convention "
                "check — assume all 10 under-table boxes sell together as ONE "
                "bulk estate lot, or skip?"),
        lot_ids=("BT-073", "BT-075", "BT-078", "BT-080"),
        value_at_stake=200.0,
        confidence_gap=0.3,
    ),
    Question(
        kind=QuestionKind.APPETITE,
        category="vintage tools",
        prompt=("Vintage Tool Sourcing (Toolbox BT-083, Wrenches BT-086): "
                "Store inventory status?"),
        lot_ids=("BT-083", "BT-086"),
        value_at_stake=150.0,
        confidence_gap=0.3,
    ),
)


def validate_cycle_questions(
    questions: tuple[Question, ...] | list[Question],
    known_lot_ids: set[str],
) -> tuple[Question, ...]:
    """Reject historical/cross-cycle questions before they enter the queue."""
    configured = tuple(questions)
    foreign_question_lots = sorted({
        lot_id
        for question in configured
        for lot_id in question.lot_ids
        if lot_id not in known_lot_ids
    })
    if foreign_question_lots:
        raise ValueError(
            "cycle questions reference lots outside this cycle: "
            + ", ".join(foreign_question_lots))
    return configured


@dataclass(frozen=True)
class IntakeStageResult:
    manifest_path: Path
    photos: tuple[dict, ...]
    telemetry: UsageTelemetry
    engine: AppraisalEngine
    triage_cache: Path
    triage_results: tuple[dict, ...]
    lot_groups: tuple
    spatial_mode: str


def run_intake_stage(
    *,
    cycle_id: str,
    listing_id: str,
    data_path: Path,
    cache_path: Path,
    force_live_vertex: bool,
) -> IntakeStageResult:
    """Load immutable source identity, triage, and evidence-gated grouping."""
    manifest_path = data_path / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"manifest not found: {manifest_path}")
    manifest_data = json.loads(manifest_path.read_text())
    photos = manifest_data["photos"]
    manifest_identity_sha256 = str(
        manifest_data.get("durable_manifest_sha256") or sha256_file(manifest_path)
    )
    print(f"[*] Ingested {len(photos)} photos from manifest "
          f"({manifest_data['captioned_photos']} captioned).")
    entries = [
        {"name": p["filename"], "uri": p["thumb_url"], "caption": p["caption"]}
        for p in photos
    ]
    parse_drop(cycle_id=cycle_id, listing_id=listing_id, entries=entries)
    telemetry = UsageTelemetry(
        cycle_id, rates_usd_per_million=rate_snapshot_usd_per_million(),
    )
    engine = AppraisalEngine(telemetry=telemetry)
    triage_cache = cache_path / "triage_results.json"
    from_cache = engine.will_use_cache(triage_cache, force_live_vertex)
    source = "cached results" if from_cache else f"LIVE Vertex AI ({TRIAGE_MODEL})"
    print(f"[*] Stage 1 Triage: {len(photos)} photos from {source}...")
    try:
        with telemetry.stage("triage.batch"):
            triage_results = engine.run_triage_batch(
                photos=photos,
                cache_path=triage_cache,
                force_refresh=force_live_vertex,
                max_workers=8,
                progress_callback=lambda done, total: (
                    print(f"    triaged {done}/{total}", end="\r")
                    if done % 25 == 0 else None
                ),
            )
    except Exception as exc:
        print(f"\n[!] Triage unavailable ({exc}); falling back to the caption heuristic.")
        triage_results = []
    verdicts = {row.get("photo_id"): row for row in triage_results}
    if triage_results:
        kept = sum(bool(row.get("worth_appraising")) for row in triage_results)
        print(f"[+] Triaged {len(triage_results)} photos; {kept} worth a closer look.")
    triaged_photos = []
    for index, photo in enumerate(photos):
        is_lot, same_lot = trusted_lot_flags(
            verdicts.get(photo["photo_id"]),
            caption=photo["caption"],
            previous_captioned=bool(index and photos[index - 1]["has_caption"]),
            index=index,
        )
        triaged_photos.append(TriagedPhoto(
            photo_id=photo["photo_id"],
            caption=photo["caption"],
            is_lot=is_lot,
            same_lot_as_previous=same_lot,
        ))
    spatial_path = data_path / "spatial_observations.json"
    spatial_mode = "walk-only"
    if spatial_path.is_file():
        observations = load_observations(
            spatial_path,
            expected_photo_ids={str(photo["photo_id"]) for photo in photos},
            expected_manifest_sha256=manifest_identity_sha256,
        )
        by_id = {observation.photo_id: observation for observation in observations}
        tagged = []
        for triaged, photo in zip(triaged_photos, photos, strict=True):
            observation = by_id[triaged.photo_id]
            tagged.append(SpatiallyTaggedPhoto(
                photo_id=triaged.photo_id,
                caption=triaged.caption,
                summary=observation.summary,
                is_lot=triaged.is_lot,
                same_lot_as_previous=triaged.same_lot_as_previous,
                surface=observation.surface,
                zone=observation.zone,
                margin_neighbors=observation.margin_neighbors,
            ))
        triaged_photos = apply_trajectory(tagged)
        spatial_mode = "validated-listing-graph"
    photo_by_seq = {photo["sequence"]: photo["photo_id"] for photo in photos}
    sequences = {photo["photo_id"]: photo["sequence"] for photo in photos}
    edges = load_reshoot_edges(
        cache_path / "embeddings.json",
        photo_by_seq,
        sequences,
        expected_manifest_sha256=manifest_identity_sha256,
    )

    # Identify from cached appraisals / captions only. Never call Vertex here.
    pid_to_lot = {photo["photo_id"]: f"BT-{photo['sequence']:03d}" for photo in photos}
    photo_by_id = {photo.photo_id: photo for photo in triaged_photos}
    cached_appraisals: dict = {}
    appraisal_cache_for_identify = data_path / "appraisal_results.json"
    if appraisal_cache_for_identify.exists():
        try:
            for raw in json.loads(appraisal_cache_for_identify.read_text()):
                lid = raw.get("lot_id")
                if lid:
                    cached_appraisals[lid] = raw
        except Exception as e:
            print(f"[!] Warning: Could not parse appraisal cache for puzzle identify: {e}")

    def identify(pids):
        out = {}
        for pid in pids:
            raw = cached_appraisals.get(pid_to_lot.get(pid, pid), {})
            cap = photo_by_id[pid].caption if pid in photo_by_id else ""
            out[pid] = (
                raw.get("identification") or cap,
                raw.get("category") or "unsorted",
            )
        return out

    proposal = walk_proposal_edges(triaged_photos) | {
        frozenset(e) if not isinstance(e, frozenset) else e for e in edges
    }
    clusters = puzzle_loop(
        triaged_photos,
        proposal_edges=proposal,
        identify=identify,
    )
    lot_groups = as_lot_groups(clusters)
    print(f"[+] Grouped {len(photos)} photos into {len(lot_groups)} distinct lots.")
    return IntakeStageResult(
        manifest_path=manifest_path,
        photos=tuple(photos),
        telemetry=telemetry,
        engine=engine,
        triage_cache=triage_cache,
        triage_results=tuple(triage_results),
        lot_groups=tuple(lot_groups),
        spatial_mode=spatial_mode,
    )


@dataclass(frozen=True)
class AppraisalStageResult:
    appraisal_cache: Path
    decomposition_cache: Path
    candidate_items: tuple[dict, ...]
    decomposition_candidates: tuple[dict, ...]
    raw_appraisals: tuple[dict, ...]
    grounded_rows: tuple[dict, ...]
    references: Mapping
    appraisal_by_lot: Mapping
    emitted_questions: tuple[Question, ...]


def exact_requested_rows(
    rows: list[dict], requested_ids: set[str], *, label: str,
) -> list[dict]:
    """Return one row per requested id; reject missing/duplicate batch output."""
    by_id = {}
    for row in rows:
        row_id = str(row.get("lot_id") or "")
        if row_id in by_id:
            raise RuntimeError(f"{label} returned duplicate lot id: {row_id}")
        by_id[row_id] = row
    missing = sorted(requested_ids - set(by_id))
    if missing:
        raise RuntimeError(
            f"{label} missing {len(missing)} requested lot(s): " + ", ".join(missing)
        )
    return [by_id[row_id] for row_id in sorted(requested_ids)]


def _write_exact_cache(path: Path, rows: list[dict]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(rows, indent=2))
    temporary.replace(path)


def run_appraisal_stage(
    *,
    intake: IntakeStageResult,
    cache_path: Path,
    references: Mapping,
    standing_rules: tuple[StandingRule, ...],
    force_live_vertex: bool,
    enable_grounded_pricing: bool,
    pricing_min_fit: float,
    pricing_workers: int,
    pricing_engine_factory=None,
) -> AppraisalStageResult:
    """Run container enrichment, appraisal, and optional cited pricing."""
    photos = list(intake.photos)
    lot_groups = list(intake.lot_groups)
    triage_results = list(intake.triage_results)
    refs = dict(references)
    appraisal_cache = cache_path / "appraisal_results.json"
    primary_ids = {group.primary_photo_id for group in lot_groups}
    lot_photos = [photo for photo in photos if photo["photo_id"] in primary_ids]
    candidate_items = select_appraisal_candidates(
        triage_results, lot_photos, always_include=set(refs), category_hints=refs,
    )
    decomposition_candidates = select_decomposition_candidates(
        triage_results,
        lot_photos,
        appraised_lot_ids={candidate["lot_id"] for candidate in candidate_items},
    )
    decomposition_cache = cache_path / "decomposition_results.json"
    if decomposition_candidates:
        from_decomposition_cache = intake.engine.will_use_cache(
            decomposition_cache,
            force_live_vertex,
            required_ids={candidate["lot_id"] for candidate in decomposition_candidates},
        )
        source = ("cached results" if from_decomposition_cache
                  else "LIVE Vertex AI spatial isolation")
        print(f"[*] Container Decomposition: {len(decomposition_candidates)} lots "
              f"from {source}...")
    # RULING (pipeline lane, 2026-08-29, asked for by the seal-path brief):
    # all-or-nothing stays. A sealed cycle must be one model vintage — letting
    # cached Aug-20 rows sit beside live fill-ins is the "partial run wearing
    # a cache" failure will_use_cache's own docstring names, and the seal's
    # value is coherence, not spend. The sanctioned resumable layer is
    # grounded pricing, whose per-lot fingerprint-gated cache makes a stopped
    # pricing run cheap to resume without mixing appraisal vintages.
    from_cache = intake.engine.will_use_cache(
        appraisal_cache,
        force_live_vertex,
        required_ids={candidate["lot_id"] for candidate in candidate_items},
    )
    source = "cached results" if from_cache else "LIVE Vertex AI (gemini-3.6-flash)"
    grounded_pipeline = None
    if enable_grounded_pricing:
        grounded_cache = cache_path / "grounded_prices.json"
        engine_factory = (
            pricing_engine_factory
            if pricing_engine_factory is not None
            else lambda: AppraisalEngine(telemetry=intake.telemetry)
        )
        print(f"[*] Grounded Pricing: appraisals enter at fit >= {pricing_min_fit:.2f} "
              f"as they finish, with {pricing_workers} workers and {MIN_CALLS} "
              "independent samples each...")
        grounded_pipeline = GroundedPricingPipeline(
            grounded_cache,
            min_fit=pricing_min_fit,
            workers=pricing_workers,
            excluded_lot_ids=set(refs),
            progress_callback=lambda done, _total: (
                print(f"    grounded {done} completed", end="\r")
                if done % 5 == 0 else None
            ),
            engine_factory=engine_factory,
        )
    print(f"\n[*] Stage 2 Appraisal: {len(candidate_items)} candidate lots from "
          f"{source}; each starts as soon as its own inputs are ready...")
    try:
        with intake.telemetry.stage("appraisal_enrichment.batch"):
            raw_appraisals, decompositions = (
                intake.engine.run_enrichment_appraisal_pipeline(
                    candidate_items,
                    decomposition_candidates=decomposition_candidates,
                    standing_rules=standing_rules,
                    appraisal_cache_path=appraisal_cache,
                    decomposition_cache_path=decomposition_cache,
                    force_refresh=force_live_vertex,
                    appraisal_workers=4,
                    decomposition_workers=4,
                    appraisal_result_callback=(
                        grounded_pipeline.submit if grounded_pipeline else None
                    ),
                )
            )
            grounded_rows = grounded_pipeline.finish() if grounded_pipeline else []
    except Exception:
        if grounded_pipeline:
            grounded_pipeline.shutdown()
        raise
    requested_appraisals = {row["lot_id"] for row in candidate_items}
    raw_appraisals = exact_requested_rows(
        raw_appraisals, requested_appraisals, label="appraisal batch",
    )
    if (not appraisal_cache.is_file()
            or len(raw_appraisals) != len(json.loads(appraisal_cache.read_text()))):
        _write_exact_cache(appraisal_cache, raw_appraisals)
    requested_decompositions = {row["lot_id"] for row in decomposition_candidates}
    if requested_decompositions:
        decompositions = exact_requested_rows(
            decompositions, requested_decompositions, label="decomposition batch",
        )
        if (not decomposition_cache.is_file()
                or len(decompositions) != len(json.loads(decomposition_cache.read_text()))):
            _write_exact_cache(decomposition_cache, decompositions)
    print(f"[{'~' if from_cache else '✓'}] Retrieved {len(raw_appraisals)} "
          f"structured appraisals from {source}."
          + ("" if from_cache else f" Written to {appraisal_cache}."))
    if grounded_pipeline:
        grounded_refs = grounded_reference_comps(grounded_rows)
        refs = {**grounded_refs, **refs}
        print(f"[+] Grounded Pricing: {len(grounded_refs)}/{len(grounded_rows)} "
              "candidate lots have usable cited sold comps.")
    appraisal_by_lot = {}
    emitted_questions = []
    for raw in raw_appraisals:
        lot_id = raw.get("lot_id")
        category_hint = refs.get(lot_id, {}).get("cat")
        appraisal, questions = intake.engine.parse_appraisal_to_domain(
            raw, category_override=category_hint,
        )
        appraisal_by_lot[appraisal.lot_id] = (appraisal, raw)
        emitted_questions.extend(questions)
    return AppraisalStageResult(
        appraisal_cache=appraisal_cache,
        decomposition_cache=decomposition_cache,
        candidate_items=tuple(candidate_items),
        decomposition_candidates=tuple(decomposition_candidates),
        raw_appraisals=tuple(raw_appraisals),
        grounded_rows=tuple(grounded_rows),
        references=MappingProxyType(refs),
        appraisal_by_lot=MappingProxyType(appraisal_by_lot),
        emitted_questions=tuple(emitted_questions),
    )


@dataclass(frozen=True)
class DecisionStageResult:
    lots: tuple[Lot, ...]
    decisions: tuple[Decision, ...]
    summary: object
    queue: object
    captions: Mapping[str, str]
    references: Mapping


def run_decision_stage(
    *,
    intake: IntakeStageResult,
    appraisal: AppraisalStageResult,
    approvals: Mapping,
    standing_rules: tuple[StandingRule, ...],
    cycle_questions: tuple[Question, ...],
    budget_cap: float,
    auto_send_threshold: float,
) -> DecisionStageResult:
    """Apply questions, mechanics, cited comps, and deterministic allocation."""
    photos = list(intake.photos)
    lot_groups = list(intake.lot_groups)
    refs = dict(appraisal.references)
    appraisal_by_lot = dict(appraisal.appraisal_by_lot)
    grounded_by_lot = {row["lot_id"]: row for row in appraisal.grounded_rows}
    known_lot_ids = {
        f"BT-{next(photo for photo in photos if photo['photo_id'] == group.primary_photo_id)['sequence']:03d}"
        for group in lot_groups
    }
    configured_questions = validate_cycle_questions(cycle_questions, known_lot_ids)
    queue_result = build_queue(
        [*configured_questions, *appraisal.emitted_questions],
        standing_rules,
        cap=12,
    )
    print(f"[+] Question Queue: {len(queue_result.asked)} asked, "
          f"{len(queue_result.auto_answered)} auto-answered from standing rules.")
    lots = []
    decisions = []
    captions_map = {}
    for group in lot_groups:
        primary = next(
            photo for photo in photos if photo["photo_id"] == group.primary_photo_id
        )
        caption = primary["caption"]
        lot_id = f"BT-{primary['sequence']:03d}"
        captions_map[lot_id] = caption
        app_pair = appraisal_by_lot.get(lot_id)
        raw_app = app_pair[1] if app_pair else {}
        if lot_id in refs:
            comp_info = refs[lot_id]
            fit, penalty = operator_lot_inputs(lot_id, raw_app, dict(approvals))
            category = comp_info["cat"]
            identification = append_visible_contents(
                raw_app.get("identification", comp_info["desc"]),
                raw_app.get("container_decomposition"),
            )
            per_unit = is_choice_lot(identification, caption)
            contents = raw_app.get("contents")
            if raw_app.get("is_container") and contents:
                identification = f"{identification}: {', '.join(contents)}"
            comp_estimate, comp_info, upside_note = comp_from_reference(
                comp_info, raw_app,
            )
            refs[lot_id] = comp_info
            if upside_note:
                identification = f"{identification}. {upside_note}"
            ruling = (approvals.get(lot_id) or {}).get("ruling")
            if ruling:
                mechanic, units, wanted = mechanic_from_ruling(ruling)
            elif per_unit:
                mechanic, units, wanted = BidMechanic.CHOICE, 1, 1
            else:
                mechanic, units, wanted = BidMechanic.STRAIGHT, 1, None
            lot = Lot(
                lot_id=lot_id,
                caption=identification,
                category=category,
                fit_score=fit,
                condition_penalty=penalty,
                comp=comp_estimate,
                mechanic=mechanic,
                unit_count=units,
                units_wanted=None,
            )
            if wanted is not None and mechanic is not BidMechanic.UNKNOWN:
                lot = elect(lot, wanted)
            decision = apply_operator_cap(price_lot(lot), dict(approvals))
        else:
            identification = append_visible_contents(
                raw_app.get("identification")
                or caption
                or f"Uncaptioned lot (Photo #{primary['sequence']})",
                raw_app.get("container_decomposition"),
            )
            fit, penalty = (
                operator_lot_inputs(lot_id, raw_app, dict(approvals))
                if app_pair else (0.20, 0.10)
            )
            category = raw_app.get("category") or "general estate"
            per_unit = is_choice_lot(identification, caption)
            mechanic, units, wanted = (
                (BidMechanic.CHOICE, 1, 1)
                if per_unit else (BidMechanic.STRAIGHT, 1, None)
            )
            lot = Lot(
                lot_id=lot_id,
                caption=identification,
                category=category,
                fit_score=fit,
                condition_penalty=penalty,
                comp=CompEstimate(
                    low=None,
                    high=None,
                    source_count=0,
                    confidence=BidConfidence.NONE,
                ),
                mechanic=mechanic,
                unit_count=units,
                units_wanted=wanted,
            )
            decision = price_lot(lot)
            if decision.needs_deep_comps:
                decision = replace(
                    decision,
                    reason=grounded_status_reason(grounded_by_lot.get(lot_id)),
                )
            decision = apply_operator_cap(decision, dict(approvals))
        lots.append(lot)
        decisions.append(decision)
    allocated = allocate(
        decisions,
        budget_cap=budget_cap,
        auto_send_threshold=auto_send_threshold,
    )
    summary = summarize(allocated)
    print("\n" + "=" * 80)
    print("ALLOCATION & SOURCING SUMMARY:")
    print(f"  Total Lots:             {summary.total_lots}")
    print(f"  Allocated Lots:         {summary.allocated}")
    print(f"  Auto-Send (<={auto_send_threshold:.2f}):   {summary.auto_send}")
    print(f"  Needs Owner Approval:   {summary.needs_approval}")
    print(f"  Committed Max Bids:     ${summary.committed_max:,.2f}")
    print(f"  Committed All-In Cost:  ${summary.committed_all_in:,.2f} "
          f"(w/ {ABSENTEE_FEE:.0%} absentee fee)")
    print("=" * 80)
    return DecisionStageResult(
        lots=tuple(lots),
        decisions=tuple(allocated),
        summary=summary,
        queue=queue_result,
        captions=MappingProxyType(captions_map),
        references=MappingProxyType(refs),
    )


def pipeline_cache_dir(data_path: Path, output_dir: Path | str | None) -> Path:
    """Caches are cycle inputs living beside the manifest. Routing outputs
    elsewhere must never move the cache lookup — coupling the two is what sent
    the 2026-08-29 rerun to Vertex at full corpus price while every cache sat
    unread in the data dir."""
    return Path(data_path)


def pipeline_state_path(
    data_path: Path, output_path: Path, explicit_output: bool,
) -> Path:
    """Where the provenance-rich state seals. The canonical (no explicit
    output) run must land it where media/video_manifest.json points — the
    facts collector reads only there — while an explicit-output run (the
    cloud worker) publishes it from its own output_dir."""
    return (output_path if explicit_output else Path(data_path)) / "pipeline_state.json"


def email_artifact_path(
    data_path: Path, output_path: Path, explicit_output: bool,
) -> Path:
    return (
        output_path / "absentee_bid_email.txt"
        if explicit_output else data_path.parent / "aug22_absentee_bid_email.txt"
    )


def sheet_artifact_path(
    data_path: Path, output_path: Path, explicit_output: bool,
) -> Path:
    return (
        output_path / "bid_sheet.xlsx"
        if explicit_output else data_path.parent / "BlueToad_2026-08-22_BidSheet.xlsx"
    )


def write_email_artifact(
    *,
    output_path: Path,
    data_path: Path,
    explicit_output: bool,
    email_to: str,
    auction_title: str,
    auction_date: str,
    venue: str,
    deadline: str,
    decision_stage: DecisionStageResult,
) -> Path:
    """Write the clerk draft from the exact decision stage."""
    path = email_artifact_path(data_path, output_path, explicit_output)
    path.write_text(compile_absentee_email(
        to=email_to,
        subject=(f"Absentee Bids - {auction_title} - {auction_date} "
                 "(Bidder: Richmond General)"),
        auction_date=auction_date,
        venue=venue,
        deadline=deadline,
        lots=list(decision_stage.lots),
        decisions=list(decision_stage.decisions),
    ))
    print(f"\n[✓] Compiled absentee bid email draft: {path}")
    return path


def write_bid_sheet_artifact(
    *,
    output_path: Path,
    data_path: Path,
    explicit_output: bool,
    decision_stage: DecisionStageResult,
) -> Path:
    """Write the workbook from the same lots, decisions, and evidence map."""
    path = sheet_artifact_path(data_path, output_path, explicit_output)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Bid Sheet"
    header_fill = PatternFill(
        start_color="1F497D", end_color="1F497D", fill_type="solid",
    )
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    approved_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid",
    )
    headers = [
        "Lot ID", "Category", "Description", "Est Resale ($)",
        "Start Bid ($)", "Max Bid ($)", "All-In ($)", "Units",
        "Committed Max ($)", "Committed All-In ($)", "Status",
        "Price Evidence", "Citations",
    ]
    sheet.append(headers)
    for column in range(1, len(headers) + 1):
        sheet.cell(1, column).fill = header_fill
        sheet.cell(1, column).font = header_font
    lots = {lot.lot_id: lot for lot in decision_stage.lots}
    references = decision_stage.references
    approved = [
        decision for decision in decision_stage.decisions
        if decision.allocated and decision.max_bid
    ]
    for row_index, decision in enumerate(approved, 2):
        lot = lots[decision.lot_id]
        evidence = references.get(decision.lot_id, {})
        estimated = (
            f"${lot.comp.low:.0f}-${lot.comp.high:.0f}" if lot.comp.low else "N/A"
        )
        sheet.append([
            decision.lot_id,
            decision.category,
            lot.caption,
            estimated,
            opening_bid(decision.max_bid),
            decision.max_bid,
            decision.all_in,
            units_committed(
                decision.mechanic, decision.unit_count, decision.units_wanted,
            ),
            decision.committed_max,
            decision.committed_all_in,
            "AUTO-SEND" if decision.auto_send else "APPROVED",
            evidence.get("provenance", "operator_reference"),
            "\n".join(evidence.get("citations") or []),
        ])
        sheet.cell(row_index, 11).fill = approved_fill
    summary = decision_stage.summary
    sheet.append([
        "", "", "", "", "", "", "", "TOTAL",
        summary.committed_max, summary.committed_all_in, "", "", "",
    ])
    for column in range(8, 11):
        sheet.cell(sheet.max_row, column).font = Font(
            name="Arial", size=11, bold=True,
        )
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        letter = get_column_letter(column[0].column)
        sheet.column_dimensions[letter].width = min(max(width + 3, 12), 80)
    workbook.save(path)
    print(f"[✓] Saved bid sheet: {path}")
    return path


def write_pipeline_state_artifact(
    *,
    cycle_id: str,
    listing_id: str,
    auction_title: str,
    auction_date: str,
    auction_timezone: str,
    auction_deadline: str,
    venue: str,
    budget_cap: float,
    auto_send_threshold: float,
    data_path: Path,
    output_path: Path,
    explicit_output: bool,
    intake: IntakeStageResult,
    appraisal: AppraisalStageResult,
    decision_stage: DecisionStageResult,
    standing_rules: tuple[StandingRule, ...],
    enable_grounded_pricing: bool,
    email_path: Path,
    bid_sheet_path: Path,
) -> Path:
    """Write provenance-rich state/telemetry and the normalized output manifest."""
    lots = {lot.lot_id: lot for lot in decision_stage.lots}
    references = decision_stage.references
    decision_rows = []
    for decision in decision_stage.decisions:
        lot = lots[decision.lot_id]
        evidence = references.get(decision.lot_id, {})
        decision_rows.append({
            "lot_id": decision.lot_id,
            "category": decision.category,
            "priority": decision.priority.value,
            "allocated": decision.allocated,
            "auto_send": decision.auto_send,
            "speculative": decision.speculative,
            "max_bid": decision.max_bid,
            "all_in": decision.all_in,
            "committed_max": decision.committed_max,
            "committed_all_in": decision.committed_all_in,
            "mechanic": decision.mechanic.value,
            "unit_count": decision.unit_count,
            "units_wanted": decision.units_wanted,
            "needs_human_pricing": decision.needs_human_pricing,
            "needs_mechanic_ruling": decision.needs_mechanic_ruling,
            "reason": decision.reason,
            "comp": {
                "low": lot.comp.low,
                "high": lot.comp.high,
                "source_count": lot.comp.source_count,
                "confidence": lot.comp.confidence.value,
                "provenance": (
                    evidence.get("provenance", "operator_reference")
                    if decision.lot_id in references else None
                ),
                "citations": list(evidence.get("citations") or []),
            },
        })
    decomposition_rows = (
        json.loads(appraisal.decomposition_cache.read_text())
        if appraisal.decomposition_cache.is_file() else []
    )
    absorption_path = data_path / "absorption_evidence.json"
    absorption_records = (
        load_absorption_evidence(absorption_path) if absorption_path.is_file() else []
    )
    absorption_revision = sha256_file(absorption_path) if absorption_records else None
    absorption_output = output_path / "absorption_evidence.json"
    if explicit_output and absorption_records:
        shutil.copy2(absorption_path, absorption_output)
    usage_file = output_path / "usage_telemetry.json"
    telemetry_payload = intake.telemetry.aggregate()
    photos = list(intake.photos)
    state_path = pipeline_state_path(data_path, output_path, explicit_output)
    state_path.write_text(json.dumps({
        "cycle_id": cycle_id,
        "listing_id": listing_id,
        "auction": {
            "title": auction_title,
            "date": auction_date,
            "timezone": auction_timezone,
            "deadline": auction_deadline,
            "venue": venue,
        },
        "budget_cap": budget_cap,
        "auto_send_threshold": auto_send_threshold,
        "summary": asdict(decision_stage.summary),
        "decisions": decision_rows,
        "queue": decision_stage.queue.accounting(),
        "standing_rules": [
            {
                "kind": rule.kind.value,
                "category": rule.category,
                "answer": rule.answer,
                "learned_cycle": rule.learned_cycle,
            }
            for rule in standing_rules
        ],
        "models": {
            "triage": TRIAGE_MODEL,
            "appraisal": APPRAISAL_MODEL,
            "grounded_pricing": bool(enable_grounded_pricing),
        },
        "performance_and_cost": {
            "planning_estimate": {
                "kind": "estimate",
                "rates_usd_per_million": rate_snapshot_usd_per_million(),
                "cost_usd": estimate_cost_usd(
                    len(photos),
                    len(appraisal.candidate_items),
                    len(appraisal.decomposition_candidates),
                ),
            },
            "measured": telemetry_payload["summary"],
        },
        "coverage": {
            "source_photo_ids": sorted(str(photo["photo_id"]) for photo in photos),
            "triage_success_ids": sorted(
                str(row.get("photo_id")) for row in intake.triage_results
            ),
            "appraisal_requested_ids": sorted(
                str(row["lot_id"]) for row in appraisal.candidate_items
            ),
            "appraisal_success_ids": sorted(
                str(row.get("lot_id")) for row in appraisal.raw_appraisals
            ),
            "grounded_attempt_ids": sorted(
                str(row.get("lot_id")) for row in appraisal.grounded_rows
            ),
            "decomposition_requested_ids": sorted(
                str(row["lot_id"]) for row in appraisal.decomposition_candidates
            ),
            "decomposition_success_ids": sorted(
                str(row.get("lot_id")) for row in decomposition_rows
            ),
        },
        "spatial": {
            "mode": intake.spatial_mode,
            "observations": len(photos) if intake.spatial_mode != "walk-only" else 0,
            "unknown_zone_default": Zone.UNKNOWN.value,
        },
        "external_evidence": {
            "absorption": {
                "status": "verified" if absorption_records else "unavailable",
                "revision_sha256": absorption_revision,
                "records": [record.as_dict() for record in absorption_records],
                "metric": "sold_units_last_365_days / active_listings_now",
                "days_on_market_used": False,
            },
        },
        "approved_lots_count": sum(
            decision.allocated and not decision.speculative
            for decision in decision_stage.decisions
        ),
        "total_lots_count": len(intake.lot_groups),
        "photos_count": len(photos),
        "grounded_pricing": {
            "enabled": enable_grounded_pricing,
            "attempted": len(appraisal.grounded_rows),
            "usable": sum(bool(row.get("usable")) for row in appraisal.grounded_rows),
        },
        "source_manifest": "manifest.json",
        "artifacts": {
            "email": email_path.name,
            "bid_sheet": bid_sheet_path.name,
            "triage": intake.triage_cache.name,
            "appraisals": appraisal.appraisal_cache.name,
            "grounded_prices": (
                "grounded_prices.json" if enable_grounded_pricing else None
            ),
            "usage_telemetry": usage_file.name,
            "absorption_evidence": (
                absorption_output.name
                if explicit_output and absorption_records else None
            ),
        },
    }, indent=2))
    intake.telemetry.write(usage_file)
    print(f"[✓] Saved pipeline state snapshot: {state_path}")
    published_manifest = output_path / "manifest.json"
    if explicit_output and intake.manifest_path.resolve() != published_manifest.resolve():
        shutil.copy2(intake.manifest_path, published_manifest)
    return state_path


def run_pipeline(
    cycle_id: str = "2026-08-22",
    listing_id: str = "4160518",
    data_dir: str = "data/aug22_gallery_4160518",
    output_dir: str | None = None,
    budget_cap: float = 600.0,
    auto_send_threshold: float = 35.0,
    force_live_vertex: bool = False,
    auction_title: str | None = None,
    auction_date: str | None = None,
    auction_timezone: str | None = None,
    auction_deadline: str | None = None,
    venue: str | None = None,
    email_to: str = "info@bluetoadauctions.com",
    reference_comps: dict | None = None,
    operator_approved: dict | None = None,
    standing_rules: tuple[StandingRule, ...] | list[StandingRule] | None = None,
    cycle_questions: tuple[Question, ...] | list[Question] | None = None,
    enable_grounded_pricing: bool = False,
    pricing_min_fit: float = 0.70,
    pricing_workers: int = 6,
    pricing_engine_factory=None,
):
    execution_metadata = {
        "auction_title": auction_title,
        "auction_date": auction_date,
        "auction_timezone": auction_timezone,
        "auction_deadline": auction_deadline,
        "venue": venue,
    }
    missing_metadata = [name for name, value in execution_metadata.items()
                        if not str(value or "").strip()]
    if missing_metadata:
        raise ValueError(
            "missing cycle execution metadata: " + ", ".join(missing_metadata))

    print("\n" + "=" * 80)
    print("BLUE TOAD FLEET — LIVE VERTEX AI SOURCING PIPELINE")
    print(f"Cycle: {cycle_id} | Listing ID: {listing_id} | Budget Cap: ${budget_cap:.2f}")
    print("=" * 80)

    # 1. Intake & Spatial Grouping
    data_path = Path(data_dir)
    output_path = Path(output_dir) if output_dir else data_path.parent
    cache_path = pipeline_cache_dir(data_path, output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    cache_path.mkdir(parents=True, exist_ok=True)
    refs = {} if reference_comps is None else reference_comps
    approvals = {} if operator_approved is None else operator_approved
    configured_rules = tuple(standing_rules or ())
    configured_questions = tuple(cycle_questions or ())

    intake = run_intake_stage(
        cycle_id=cycle_id,
        listing_id=listing_id,
        data_path=data_path,
        cache_path=cache_path,
        force_live_vertex=force_live_vertex,
    )

    appraisal = run_appraisal_stage(
        intake=intake,
        cache_path=cache_path,
        references=refs,
        standing_rules=configured_rules,
        force_live_vertex=force_live_vertex,
        enable_grounded_pricing=enable_grounded_pricing,
        pricing_min_fit=pricing_min_fit,
        pricing_workers=pricing_workers,
        pricing_engine_factory=pricing_engine_factory,
    )

    decision_stage = run_decision_stage(
        intake=intake,
        appraisal=appraisal,
        approvals=approvals,
        standing_rules=configured_rules,
        cycle_questions=configured_questions,
        budget_cap=budget_cap,
        auto_send_threshold=auto_send_threshold,
    )

    resolved_auction_date = str(auction_date)
    email_draft_path = write_email_artifact(
        output_path=output_path,
        data_path=data_path,
        explicit_output=bool(output_dir),
        email_to=email_to,
        auction_title=str(auction_title),
        auction_date=resolved_auction_date,
        venue=venue,
        deadline=auction_deadline,
        decision_stage=decision_stage,
    )
    out_excel = write_bid_sheet_artifact(
        output_path=output_path,
        data_path=data_path,
        explicit_output=bool(output_dir),
        decision_stage=decision_stage,
    )

    pipeline_state_file = write_pipeline_state_artifact(
        cycle_id=cycle_id,
        listing_id=listing_id,
        auction_title=str(auction_title),
        auction_date=resolved_auction_date,
        auction_timezone=str(auction_timezone),
        auction_deadline=str(auction_deadline),
        venue=str(venue),
        budget_cap=budget_cap,
        auto_send_threshold=auto_send_threshold,
        data_path=data_path,
        output_path=output_path,
        explicit_output=bool(output_dir),
        intake=intake,
        appraisal=appraisal,
        decision_stage=decision_stage,
        standing_rules=configured_rules,
        enable_grounded_pricing=enable_grounded_pricing,
        email_path=email_draft_path,
        bid_sheet_path=out_excel,
    )

    return PipelineResult(
        photos=intake.photos,
        lot_groups=intake.lot_groups,
        lots=decision_stage.lots,
        decisions=decision_stage.decisions,
        summary=decision_stage.summary,
        queue=decision_stage.queue,
        captions=decision_stage.captions,
        pipeline_state_path=pipeline_state_file,
        bid_sheet_path=out_excel,
        email_path=email_draft_path,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--live", action="store_true", help="bypass model caches")
    parser.add_argument("--data-dir", type=Path,
                        default=Path("data/aug22_gallery_4160518"))
    # No default output dir: the canonical invocation seals state beside the
    # caches (where the facts collector reads) and keeps the money artifacts
    # on their protected historical paths. Pass --output-dir only to stage a
    # publishable copy elsewhere, as the cloud worker does.
    parser.add_argument("--output-dir", type=Path, default=None)
    args = parser.parse_args(argv)
    try:
        execute_pipeline(PipelineConfig(
            cycle_id="2026-08-22",
            listing_id="4160518",
            data_dir=args.data_dir,
            output_dir=args.output_dir,
            budget_cap=600.0,
            auto_send_threshold=35.0,
            force_live_vertex=args.live,
            auction_title="Blue Toad Auctions Estate Sale",
            auction_date="2026-08-22",
            auction_timezone="America/Chicago",
            auction_deadline="2026-08-21T16:00:00-05:00",
            venue="200 Elizabeth Lane, Genoa City, WI",
            email_to="info@bluetoadauctions.com",
            reference_comps=MappingProxyType(REFERENCE_COMPS),
            operator_approved=MappingProxyType(OPERATOR_APPROVED),
            standing_rules=tuple(DEFAULT_STANDING_RULES),
            cycle_questions=tuple(AUG22_DOMAIN_QUESTIONS),
        ))
    except (OSError, RuntimeError, ValueError) as exc:
        print(f"pipeline failed: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
