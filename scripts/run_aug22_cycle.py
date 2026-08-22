#!/usr/bin/env python3
"""
scripts/run_aug22_cycle.py — Final Sealed August 22 Absentee Bid Sourcing Run.

Strict Auction Sourcing Constraints:
- Topps Cards: Selected BT-001 (Golden Era 1959-69 in top-loaders), Dropped bulk BT-284
- Costume Jewelry: BT-002, BT-087, BT-181 approved at $25.00 max bid each
- Edison Rolls: BT-041 approved at $40.00 max bid
- Fast Smalls: Princess Phone ($20), ET Nightlight ($20), Century Progress Bottle ($15),
  Lionel Set ($25), Trading Cards ($15 x 2), Handheld Games ($10)
- Excluded: Coke bottles, Marilyn plates, bulk cards BT-284, tools, dishes, autographs
- Standard Auction Bidding Increments ($5 up to $100)
- Transmission Mode: STUBBED / LOCAL DRAFT (No live emails sent)
"""

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Exact final approved lot allocations with $5 auction increments
APPROVED_BIDS = [
    # Lot ID, Sequence, Description, Category, Start Bid, Max Bid, Est Resale
    ("BT-001", 1, "Vintage Topps Baseball Cards (1959-69 Golden Era Stars)", "vintage cards", 35.00, 100.00, "$250-$500"),
    ("BT-041", 41, "Edison rolls (11-12 canisters + bare roll)", "phonograph / records", 15.00, 40.00, "$100-$160"),
    ("BT-002", 2, "Estate Costume Jewelry (Tray 12/14/16: 50-70 pcs)", "jewelry", 10.00, 25.00, "$80-$160"),
    ("BT-087", 87, "costume jewelry (Tray Lot 2)", "jewelry", 10.00, 25.00, "$80-$160"),
    ("BT-181", 181, "estate costume jewelry (Tray Lot 3)", "jewelry", 10.00, 25.00, "$80-$160"),
    ("BT-050", 50, "Lionel building set", "vintage toys", 10.00, 25.00, "$50-$120"),
    ("BT-021", 21, "princess phone", "vintage electronics", 10.00, 20.00, "$40-$100"),
    ("BT-048", 48, "ET nightlight", "vintage smalls", 10.00, 20.00, "$40-$100"),
    ("BT-235", 235, "Century Progress bottle", "advertising / bottles", 10.00, 15.00, "$30-$80"),
    ("BT-016", 16, "trading cards", "vintage cards", 10.00, 15.00, "$30-$80"),
    ("BT-030", 30, "non-sport trading cards", "vintage cards", 10.00, 15.00, "$30-$80"),
    ("BT-066", 66, "hand held video games (5 Radica/LCD units)", "vintage toys", 5.00, 10.00, "$25-$45"),
]

# The fee, the grid and the all-in formula all live in src/bidmath. This
# script previously re-declared ABSENTEE_FEE and multiplied by a literal
# 1.15 twice, which is two more places to forget when the house changes it.
from src.bidmath import ABSENTEE_FEE, all_in_cost  # noqa: E402


def main():
    total_max = sum(b[5] for b in APPROVED_BIDS)
    total_all_in = round(total_max * (1.0 + ABSENTEE_FEE), 2)

    print("\n" + "=" * 75)
    print("BLUE TOAD FLEET — FINAL APPROVED AUGUST 22 BID SHEET (STUBBED / DRAFT)")
    print("=" * 75)
    print(f"Total Approved Lots:          {len(APPROVED_BIDS)}")
    print(f"Committed Max Bids:           ${total_max:,.2f}")
    print(f"Committed All-In (w/ 15% fee):${total_all_in:,.2f}")
    print(f"Bidding Increments:           Standard $5.00 increments (up to $100)")
    print(f"Transmission Mode:            STUBBED / LOCAL REVIEW (No email sent)")
    print("=" * 75)

    print("\n[★] Final Approved Absentee Bid Schedule:")
    for lot_id, seq, desc, cat, start_bid, max_bid, est in APPROVED_BIDS:
        all_in = all_in_cost(max_bid)
        print(f"  {lot_id} | {desc:<48} | Start: ${start_bid:>5.2f} | Max: ${max_bid:>6.2f} | All-In: ${all_in:>6.2f} | Est: {est}")

    # Generate Final Sealed Absentee Bid Email Draft
    email_draft_path = Path("data/aug22_absentee_bid_email.txt")
    email_lines = [
        "TO: info@bluetoadauctions.com",
        "SUBJECT: Absentee Bids - August 22 Antique & Estate Auction (Bidder: Richmond General)",
        "DATE: Friday, August 21, 2026 (Before 8:00 PM CDT Cutoff)",
        "",
        "Blue Toad Auctions,",
        "",
        "Please register the following absentee proxy bids for the Saturday, August 22, 2026 auction",
        "at 200 Elizabeth Lane, Genoa City, WI.",
        "",
        "Bidder Info:",
        "  Name: Richmond General (Scott)",
        "  Resale Certificate: On file (Wisconsin Tax-Exempt)",
        "  Terms: 15% Absentee Buyer Fee acknowledged (Credit Card on File)",
        "",
        "-----------------------------------------------------------------------------------------",
        "ITEM DESCRIPTION                                     PHOTO REF     START ($)  MAX BID ($)",
        "-----------------------------------------------------------------------------------------",
    ]

    for lot_id, seq, desc, cat, start_bid, max_bid, est in APPROVED_BIDS:
        email_lines.append(f"{desc[:48]:<50} {lot_id:<12} ${start_bid:>8.2f} ${max_bid:>9.2f}")

    email_lines.extend([
        "-----------------------------------------------------------------------------------------",
        f"TOTAL COMMITTED PROXY BIDS: ${total_max:,.2f} (${total_all_in:,.2f} all-in w/ 15% fee)",
        "",
        "Special Instructions:",
        # NOTE: this script writes the SAME path as run_vertex_pipeline.py but
        # carries a hardcoded APPROVED_BIDS table with no bid mechanic, so it
        # cannot state a per-lot exception and running it silently reverts any
        # times-the-money instruction the pipeline wrote. Kept consistent in
        # wording; use run_vertex_pipeline.py for a real sheet.
        "  - For any OTHER 'Buyer's Choice / Times the Money' shelf lot, max quantity is 1 unit.",
        "  - Standard $5.00 bidding increments applied.",
        "  - Please confirm receipt of these absentee bids by reply email.",
        "",
        "Thank you,",
        "Richmond General",
    ])
    email_draft_path.write_text("\n".join(email_lines))
    print(f"\n[✓] Compiled Final Sealed Absentee Bid Email Draft: {email_draft_path}")

    # Generate Excel Sourcing Sheet
    out_excel = Path("data/BlueToad_2026-08-22_BidSheet.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aug 22 Bid Sheet"

    hdr_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    hdr_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    headers = ["Lot ID", "Category", "Description", "Est Resale ($)", "Start Bid ($)", "Max Bid ($)", "All-In ($)", "Status"]
    ws.append(headers)
    for c in range(1, 9):
        ws.cell(1, c).fill = hdr_fill
        ws.cell(1, c).font = hdr_font

    for idx, (lot_id, seq, desc, cat, start_bid, max_bid, est) in enumerate(APPROVED_BIDS, 2):
        all_in = all_in_cost(max_bid)
        ws.append([
            lot_id,
            cat,
            desc,
            est,
            start_bid,
            max_bid,
            all_in,
            "APPROVED",
        ])
        ws.cell(idx, 8).fill = green_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(out_excel)
    print(f"[✓] Saved updated Excel Bid Sheet: {out_excel}")


if __name__ == "__main__":
    main()
